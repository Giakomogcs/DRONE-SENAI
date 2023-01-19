from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter 
import os
from datetime import datetime
import pandas as pd


'''
# create a new XLSX workbook
wb = Workbook()
# save workbook as .xlsx file
wb.save("workbook.xlsx")

# datetime object containing current date and time
now = datetime.now()
 
print("now =", now)

# dd/mm/YY H:M:S
dt_string = now.strftime("Varredura"+"%d-%m-%Y"+"_"+"%H-%M-%S")
print("date and time =", dt_string)

os.rename('workbook.xlsx', dt_string+'.xlsx')
'''

try: 
    # datetime object containing current date and time
    back_work = pd.read_csv(r"workbook.csv")
    back_work.to_excel (r"workbook.xlsx")

    now = datetime.now()
    print("now =", now)

    # dd/mm/YY H:M:S
    dt_string = now.strftime("Varredura_"+"%d-%m-%Y"+"_"+"%H-%M-%S")
    print("date and time =", dt_string)

    os.rename('workbook.xlsx', dt_string+'.xlsx')


    # create a new XLSX workbook
    wb = Workbook()
    # save workbook as .xlsx file
    wb.save("workbook.xlsx")

    work = pd.read_excel(r"workbook.xlsx")
    work.to_csv (r"workbook.csv")

except:
    # create a new XLSX workbook
    wb = Workbook()
    # save workbook as .xlsx file
    wb.save("workbook.xlsx")

    work = pd.read_excel(r"workbook.xlsx")
    work.to_csv (r"workbook.csv")

    '''try: 
			# datetime object containing current date and time
			now = datetime.now()
			
			print("now =", now)

			# dd/mm/YY H:M:S
			dt_string = now.strftime("Varredura_"+"%d-%m-%Y"+"_"+"%H-%M-%S")
			print("date and time =", dt_string)

			os.rename('workbook.csv', dt_string+'.csv')


			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")
			root.destroy()

		except:
			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")
			root.destroy()
		'''
    