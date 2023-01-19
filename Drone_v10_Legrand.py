
import numpy as np
import cv2
import threading
''' Python program to Scan and Read a QR code'''
from pyzbar.pyzbar import decode
'''Para realizar o Json'''
import pandas as pd
import csv
from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter 
import tkinter as tk
from tkinter import *
from tkinter import messagebox as mb
from PIL import Image, ImageTk

import os
import subprocess
from datetime import datetime

import time




class Tello_2:

    def __init__(self):
        self._running = True
        #self.video = cv2.VideoCapture("udp://192.168.10.1:11111")
        #self.video = cv2.VideoCapture("udp://@0.0.0.0:11111")
        #self.video = cv2.VideoCapture("rtmp:10.84.30.32:1935/")


        ##self.video = cv2.VideoCapture("rtmp:10.84.82.101:1935/")
        self.video = cv2.VideoCapture(0)  # webcam
        
        #self.video.set(cv2.CAP_PROP_FRAME_WIDTH, 960)
        #self.video.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
        self.video.set(3,640)
        self.video.set(4,480)


    def terminate(self):
        global i,ids,past

        self._running = False
        self.video.release()
        cv2.destroyAllWindows()
        
        i = 0
        ids = []
        past = []
            


    def infos(self):
        global frame, key, ids, past, rload, cod, myData, past, LocalProps
        i = 0
        pd = False
        lc = False
        rload = False
        ids = []
        past = []
        LocalProps = []



        while self._running:

            global frame,ret   

            try:
                ret, frame = self.video.read()
                #frame = frame[0:400, 200:660]
                #frame = frame[200:1300, 600:1300]
                ##frame = frame[200:1200, 400:1000]

                if ret:

                    for barcode in decode(frame):
                        myData = barcode.data.decode('utf-8')
                        #print(myData)

                        scan(myData) #vai fazer o scan
                        
                        #print(DataPL)

                        if myData != '' and (DataPL == 'Local' or DataPL == 'Produto'): #verifica se existe o codigo lido no excel
                            mycolor = (0,255,0) 
                            

                            if ((myData not in past) and (len(ids) < 2)):

                                if ((myData not in ids) and (len(ids) < 2)):

                                    if ((DataPL == 'Produto') and (pd == False) and (len(ids) < 1)):
                                        ids.append(myData) #produto
                                        past.extend(ids) #protutos passados

                                        #print(ids)
                                        pd = True
                                        
                                    elif ((DataPL == 'Local') and (lc == False) and (len(ids) == 1)):
                                        ids.append(myData) #local
                                        ids.extend(LocalProps)
                                        past.extend(ids) #protutos passados

                                        print(ids)
                                        lc = True
                                     

                            if len(ids) == 5:
                                #past.extend(ids) #protutos passados
                                
                                post_msg()
                                #print(ids) 

                                #print(rload)
                                if rload == True:
                                    print(ids)  
                                    print(rload)    
                                    ids = []
                                    pd = False 
                                    lc = False
                                    rload = False     

                        else:
                            myoutput = 'Não Autorizado'
                            mycolor = (0,0,255)

                        if (DataPL ==''):
                            mycolor = (0,0,255)
                          

                        pts = np.array([barcode.polygon],np.int32)
                        pts = pts.reshape((-1,1,2))
                        #cv2.polylines(img,[pts],True,(176,196,222),3)
                        cv2.polylines(frame,[pts],True,(mycolor),5)

                        pts2 = barcode.rect
                        #cv2.putText(img,myData,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_SIMPLEX,0.9,(255,0,255),2)
                        #cv2.putText(img,myData,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_PLAIN,0.5,(176,196,222),1, cv2.LINE_AA)
                        cv2.putText(frame,myData,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_PLAIN,1,(mycolor),2, cv2.LINE_AA)

                    # Draw black background for textq
                    font = cv2.FONT_HERSHEY_PLAIN
                    cv2.rectangle(frame, (0, 600), (0, 120), (255, 255, 255), -1)
                    
                    #open_camera(frame)
                    #cv2.imshow('Camera Online do Drone', frame)

                    


                '''
                    cv2.waitKey(1)
                    
                    
                
                key = cv2.waitKey(1) & 0xff
                if key == 27:
                 #Apertar Esc para sair
                     
                    self.video.release()
                    cv2.destroyAllWindows()
                    root.destroy()

                    recvThread3 = threading.Thread(target=t.terminate)
                    recvThread3.stop()
                    break
                '''    

            except Exception as err:
                print(err)
    
            
            
            
def scan(data): 
    global DataPL, loc, Dloc, prod, filtroLRua, filtroLNivel, filtroLColuna, LocalProps
    DataPL = ''
    
    if myData.isdigit():
        dataMod = int(data)
        #print(type(dataMod))
    else:
        dataMod = str(data)
        #print(type(dataMod))

    try:
        try:
            filtroP = cod['Produto'] == dataMod #gera tabela true e false
            filtroPi = (cod[filtroP].iloc[0,5])
            #filtroBoolP = (filtroP[0] == True)  #saber se tem mesmo true
            DataPL = 'Produto'
            prod = data
            #print(filtroPi)

        except:
            filtroL = cod['Local'] == dataMod #gera tabela true e false
            filtroLi = (cod[filtroL].iloc[0,0])
            filtroLRua = (cod[filtroL].iloc[0,1])
            filtroLNivel = (cod[filtroL].iloc[0,2])
            filtroLColuna = (cod[filtroL].iloc[0,3])

            #filtroBoolL = (filtroL[0] == True)  #saber se tem mesmo true
            DataPL = 'Local'
            loc = data
            LocalProps = [filtroLRua,filtroLNivel,filtroLColuna]
            
            #print(filtroLi)
    except:
        DataPL = ''
        
    return DataPL




def post_msg(): #transforma o xlsx em csv para leitura e escreve o produto e local
    print('E N T R E I')
    global ids,rload,filename,LocalProps

    filename = "workbook.csv"
    header = ("Produto", "Local", "Rua", "Nivel", "Coluna")
    

    try: 
        filtered_prod = pd.read_csv("workbook.csv", usecols=["Produto"]) #gera tabela true e false
        if filtered_prod.columns == "Produto":
            with open (filename, "a", newline = "") as csvfile:  #adiciona conteudo em arquivo sem apagar
                file = csv.writer(csvfile)
                
                file.writerow(ids)
                csvfile.close()

    except: 

        with open (filename, "w", newline = "") as csvfile: #apaga tudo e escreve
            file = csv.writer(csvfile)
            file.writerow(header)
            #file.writerow(data)
            csvfile.close()

        with open (filename, "a", newline = "") as csvfile:  #adiciona conteudo em arquivo sem apagar
            file = csv.writer(csvfile)
            
            file.writerow(ids)
            csvfile.close()
    
    csv_check()
    ids = []
    LocalProps = []
    rload = True 


def open_camera():

        if(ret):
            # Convert image from one color space to other
            opencv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)

            # Capture the latest frame and transform to image
            captured_image = Image.fromarray(opencv_image)

            # Convert captured image to photoimage
            photo_image = ImageTk.PhotoImage(image=captured_image)

            # Displaying photoimage in the label
            label_widget.photo_image = photo_image

            # Configure image in the label
            label_widget.configure(image=photo_image)

            # Repeat the same process after every 10 seconds
            label_widget.after(5, open_camera)
        else:
            time.sleep(0.1)
            open_camera()


    



def explorer():
    subprocess.Popen(r'explorer /open,"C:\Users\Giovani\OneDrive - SESISENAISP - Corporativo\Área de Trabalho\IST-Projetos\legrand\Drone_le\Varreduras"')	


def csv_check():
	global ProdutoCSV, LocalCSV, RuaCSV, NivelCSV, ColunaCSV

	print ('entrei')
	# reading CSV file
	data = pd.read_csv("workbook.csv")
	
	# converting column data to list
	ProdutoCSV = data['Produto'].tolist()
	LocalCSV = data['Local'].tolist()
	RuaCSV = data['Rua'].tolist()
	NivelCSV = data['Nivel'].tolist()
	ColunaCSV = data['Coluna'].tolist()
	
	return ProdutoCSV, LocalCSV, RuaCSV, NivelCSV, ColunaCSV



def call():
	global past

	csv_check()

	res = mb.askquestion('Reset Aplication',
						'Do you really want to save as ?')
	
	if res == 'yes' :
		

		try: 
			# datetime object containing current date and time
			back_work = pd.read_csv(r"workbook.csv")
			back_work.to_excel (r"workbook.xlsx")

			now = datetime.now()
			print("now =", now)

			# dd/mm/YY H:M:S
			dt_string = now.strftime("Rua_"+RuaCSV[0]+"_"+"%d-%m-%Y"+"_"+"%H-%M-%S")
			print("date and time =", dt_string)

			os.rename('workbook.xlsx', f"Varreduras/{dt_string}'.xlsx")


			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")

			work = pd.read_excel(r"workbook.xlsx")
			work.to_csv (r"workbook.csv")
			past = []

		except:
			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")

			work = pd.read_excel(r"workbook.xlsx")
			work.to_csv (r"workbook.csv")

	else :
		mb.showinfo('Return', 'Returning to main application')

          

cod = pd.read_excel(r"codigos_GS1_v3.xlsx", engine='openpyxl')

#inicia thread de leitura da câmera
t = Tello_2()
recvThread = threading.Thread(target=t.infos)
recvThread.start()

# Main loop
if(os.path.isfile("workbook.xlsx")):
    print('file exist')
else:    
    # create a new XLSX workbook
    wb = Workbook()
    # save workbook as .xlsx file
    wb.save("workbook.xlsx")

    work = pd.read_excel(r"workbook.xlsx")
    work.to_csv (r"workbook.csv")



# Driver's code
root = Tk()
root.bind('<Escape>', lambda e: root.quit())

root.title('Camera Online do Drone')
# Create a label and display it on app
label_widget = Label(root)
label_widget.pack()

button1 = Button(root, text="Open Camera", command=open_camera)
button1.pack(side="left")

buttonSave = Button(root, text="Save as", command=call)
buttonSave.pack(side="left") 
#buttonSave.pack(side=LEFT, padx=0, pady=20) 

buttonExplorer = Button(root, text="Varreduras", command=explorer)
buttonExplorer.pack(side="right") 
#buttonExplorer.pack(side=RIGHT, padx=0, pady=20) 

root.mainloop()
