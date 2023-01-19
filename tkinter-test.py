# Python program to create
# yes/no message box

from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter 
import tkinter as tk
from tkinter import *
from tkinter import messagebox as mb
import os
from datetime import datetime
import pandas as pd


def csv_check():
	global ProdutoCSV, LocalCSV, RuaCSV, NivelCSV, ColunaCSV

	print ('entrei')
	# reading CSV file
	data = pd.read_csv("workbook.csv")
	
	# converting column data to list
	try:
		ProdutoCSV = data['Produto'].tolist()
		LocalCSV = data['Local'].tolist()
		RuaCSV = data['Rua'].tolist()
		NivelCSV = data['Nivel'].tolist()
		ColunaCSV = data['Coluna'].tolist()
		
	except:
		ProdutoCSV=[]
		LocalCSV=[]
		RuaCSV=[]
		NivelCSV=[]
		ColunaCSV=[]


	return ProdutoCSV, LocalCSV, RuaCSV, NivelCSV, ColunaCSV



def call():
	global past

	csv_check()

	res = mb.askquestion('Reset Aplication',
						'Do you really want to save as ?')
	
	if res == 'yes' :
		

		try: 
			# datetime object containing current date and time
			back_work = pd.read_csv(r"workbook.csv")
			back_work.to_excel (r"workbook.xlsx")

			now = datetime.now()
			print("now =", now)

			# dd/mm/YY H:M:S
			dt_string = now.strftime("Rua_"+RuaCSV[0]+"_"+"%d-%m-%Y"+"_"+"%H-%M-%S")
			print("date and time =", dt_string)

			os.rename('workbook.xlsx', f"Varreduras/{dt_string}'.xlsx")


			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")

			work = pd.read_excel(r"workbook.xlsx")
			work.to_csv (r"workbook.csv")
			past = []

		except:
			# create a new XLSX workbook
			wb = Workbook()
			# save workbook as .xlsx file
			wb.save("workbook.xlsx")

			work = pd.read_excel(r"workbook.xlsx")
			work.to_csv (r"workbook.csv")

	else :
		mb.showinfo('Return', 'Returning to main application')

# Driver's code
root = tk.Tk()
canvas = tk.Canvas(root,
				width = 200,
				height = 200)

canvas.pack()
b = Button(root,
		text ='Save As',
		command = call)

canvas.create_window(100, 100,
					window = b)

csv_check()
root.mainloop()
